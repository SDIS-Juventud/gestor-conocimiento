# -*- coding: utf-8 -*-
"""
Edita el Excel "Reportes Externos Subdirección Juventud 2026.xlsx" para
incluir la observación literal de Carolina (2026-06-01) sobre cómo se
reportan las políticas étnicas (Indígena, Raizal, Palenquera, Rrom) en
los tres servicios.

Cambios:
- Hoja "Mapeo general": observación nueva en filas Indígena 1.3.12,
  Raizal 4.1.12, Palenquera 6.3.15 y Rrom 6.1.8.
- Hoja "Casas de Juventud": observación nueva en fila Negra/Afro 1.3.9.

Las celdas modificadas se pintan con fondo amarillo suave para que
queden visiblemente marcadas como cambios del 2026-06-01.

Antes de correr este script, conviene tener el backup
"Reportes Externos Subdirección Juventud 2026_backup_2026-06-01.xlsx".
"""
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
RUTA = os.path.join(BASE, "ejes", "Políticas",
                    "Reportes Externos Subdirección Juventud 2026.xlsx")

OBSERVACION_NUEVA = (
    "Se solicita matriz de datos PUA (extraído de SIRBE), se especifica la "
    "necesidad de incluir datos de pertenencia étnica. El reporte para cada "
    "política se realiza filtrando por la variable de pertenencia étnica. Se "
    "diligencian matrices enviadas por DADE."
)

# Amarillo muy suave para marcar las celdas editadas el 2026-06-01.
FILL_EDITADO = PatternFill(
    start_color="FFF3B0", end_color="FFF3B0", fill_type="solid"
)


def _indice_columna(hoja, nombre_columna):
    """Devuelve el índice 1-based de la columna que se llama nombre_columna
    en la fila 1 (header). Falla si no encuentra."""
    for col in range(1, hoja.max_column + 1):
        if str(hoja.cell(row=1, column=col).value).strip() == nombre_columna:
            return col
    raise ValueError(
        f"No se encontró la columna '{nombre_columna}' en la hoja '{hoja.title}'"
    )


def _localizar_fila(hoja, col_tema, col_prod, patron_tema, patron_prod):
    """Busca la fila cuyo Tema contiene patron_tema (case-insensitive) y
    cuyo Producto comienza con patron_prod. Devuelve el número de fila
    (1-based) o None si no encuentra."""
    for fila in range(2, hoja.max_row + 1):
        tema = str(hoja.cell(row=fila, column=col_tema).value or "")
        prod = str(hoja.cell(row=fila, column=col_prod).value or "")
        if patron_tema.lower() in tema.lower() and prod.strip().startswith(patron_prod):
            return fila
    return None


def editar():
    wb = load_workbook(RUTA)

    # ----- Hoja "Mapeo general" -----
    hoja_mg = wb["Mapeo general"]
    col_tema_mg = _indice_columna(hoja_mg, "Tema")
    col_prod_mg = _indice_columna(hoja_mg, "Productos")
    col_obs_mg = _indice_columna(hoja_mg, "Observaciones")

    objetivos_mg = [
        ("Indígena", "1.3.12"),
        ("Raizal", "4.1.12"),
        ("Palenquera", "6.3.15"),
        ("Rrom", "6.1.8"),
    ]
    print("=== Hoja Mapeo general ===")
    for patron_tema, patron_prod in objetivos_mg:
        fila = _localizar_fila(hoja_mg, col_tema_mg, col_prod_mg,
                               patron_tema, patron_prod)
        if fila is None:
            print(f"  ! No se encontró fila para {patron_tema} {patron_prod}")
            continue
        celda = hoja_mg.cell(row=fila, column=col_obs_mg)
        anterior = str(celda.value or "")[:80]
        celda.value = OBSERVACION_NUEVA
        celda.fill = FILL_EDITADO
        celda.alignment = Alignment(wrap_text=True, vertical="top")
        print(f"  OK {patron_tema} {patron_prod} (fila {fila}): observacion reemplazada")
        print(f"    anterior: {anterior}...")

    # ----- Hoja "Jóvenes Con Oportunidades" — Raizal 4.1.12 -----
    hoja_jco = wb["Jóvenes Con Oportunidades"]
    col_tema_j = _indice_columna(hoja_jco, "Tema")
    col_prod_j = _indice_columna(hoja_jco, "Productos")
    col_obs_j = _indice_columna(hoja_jco, "Observaciones")

    print("\n=== Hoja Jóvenes Con Oportunidades ===")
    fila = _localizar_fila(hoja_jco, col_tema_j, col_prod_j,
                            "Raizal", "4.1.12")
    if fila is None:
        print("  ! No se encontró fila Raizal 4.1.12 en JCO")
    else:
        celda = hoja_jco.cell(row=fila, column=col_obs_j)
        anterior = str(celda.value or "")[:80]
        celda.value = OBSERVACION_NUEVA
        celda.fill = FILL_EDITADO
        celda.alignment = Alignment(wrap_text=True, vertical="top")
        print(f"  OK Raizal 4.1.12 (fila {fila}): observacion reemplazada")
        print(f"    anterior: {anterior}...")

    # ----- Hoja "Casas de Juventud" -----
    hoja_casas = wb["Casas de Juventud"]
    col_tema_c = _indice_columna(hoja_casas, "Tema")
    col_prod_c = _indice_columna(hoja_casas, "Productos")
    col_obs_c = _indice_columna(hoja_casas, "Observaciones")

    print("\n=== Hoja Casas de Juventud ===")
    fila = _localizar_fila(hoja_casas, col_tema_c, col_prod_c,
                            "Negra", "1.3.9")
    if fila is None:
        print("  ! No se encontró fila Negra/Afro 1.3.9 en Casas")
    else:
        celda = hoja_casas.cell(row=fila, column=col_obs_c)
        anterior = str(celda.value or "")[:80]
        celda.value = OBSERVACION_NUEVA
        celda.fill = FILL_EDITADO
        celda.alignment = Alignment(wrap_text=True, vertical="top")
        print(f"  OK Negra/Afro 1.3.9 (fila {fila}): observacion reemplazada")
        print(f"    anterior: {anterior}...")

    wb.save(RUTA)
    print(f"\nArchivo guardado: {RUTA}")


if __name__ == "__main__":
    editar()
